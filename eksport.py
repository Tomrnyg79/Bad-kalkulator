"""
Eksportfunksjoner for PDF, Excel og e-post.
Kun post og totalpris – ingen enhetspriser.
"""

import io
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import pathlib
from priser import FIRMA, MVA_SATS

_LOGO = pathlib.Path(__file__).parent / "unnamed.jpg"


def fmt(tall):
    return f"{tall:,.0f}".replace(",", " ")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _pdf_seksjon(pdf, tittel, poster, kol_b):
    """poster: liste med (navn, mengde, enhet, enhetspris, total) tupler."""
    if not poster:
        return
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, tittel, new_x="LMARGIN", new_y="NEXT")

    overskrifter = ["Post", "Mengde", "Enhet", "Enh.pris", "Sum kr"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(50, 50, 50)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(overskrifter):
        pdf.cell(kol_b[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 9)
    skravur = False
    for navn, mengde, enhet, enhetspris, total in poster:
        pdf.set_fill_color(245, 245, 245) if skravur else pdf.set_fill_color(255, 255, 255)
        pdf.cell(kol_b[0], 6, navn, border=1, fill=True)
        pdf.cell(kol_b[1], 6, f"{mengde:.2f}", border=1, fill=True, align="R")
        pdf.cell(kol_b[2], 6, enhet, border=1, fill=True, align="C")
        pdf.cell(kol_b[3], 6, fmt(enhetspris), border=1, fill=True, align="R")
        pdf.cell(kol_b[4], 6, fmt(total), border=1, fill=True, align="R")
        pdf.ln()
        skravur = not skravur

    seksjon_sum = sum(t for _, _, _, _, t in poster)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(kol_b[0] + kol_b[1] + kol_b[2] + kol_b[3], 7, f"Sum {tittel.lower()}", border=1, fill=True)
    pdf.cell(kol_b[4], 7, fmt(seksjon_sum), border=1, fill=True, align="R")
    pdf.ln()
    pdf.ln(4)


def generer_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Firmaheader med logo
    pdf.set_font("Helvetica", "B", 20)
    if _LOGO.exists():
        pdf.image(str(_LOGO), x=10, y=10, w=50)
        pdf.set_y(10)
        pdf.cell(55)  # hopp forbi logoen
    pdf.cell(0, 10, FIRMA["navn"], new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_x(65)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 4, f"Org.nr: {FIRMA['orgnr']}  |  Tlf: {FIRMA['telefon']}", new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_x(65)
    pdf.cell(0, 4, f"{FIRMA['adresse']}  |  {FIRMA['epost']}", new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_y(max(pdf.get_y(), 40))  # sørg for plass under logoen

    pdf.ln(3)
    pdf.set_draw_color(180, 180, 180)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(6)

    # Tittel
    er_manuell = data.get("kalkyle_type") == "manuell"
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, "KALKYLE" if er_manuell else "BADEROMS KALKYLE", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Prosjektinfo
    if er_manuell:
        felter = [
            ("Prosjekt:", data.get("adresse", "")),
            ("Dato:", data.get("dato", "")),
        ]
    else:
        felter = [
            ("Prosjekt:", data.get("adresse", "")),
            ("Dato:", data.get("dato", "")),
            ("Gulvareal:", f"{data.get('gulvareal', 0):.2f} m\u00b2"),
            ("Veggareal:", f"{data.get('veggareal', 0):.2f} m\u00b2"),
        ]
    for etikett, verdi in felter:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(28, 6, etikett)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, verdi, new_x="LMARGIN", new_y="NEXT")

    pdf.ln(6)

    kol_b = [80, 22, 18, 32, 38]  # Post, Mengde, Enhet, Enh.pris, Sum = 190

    if er_manuell:
        _pdf_seksjon(pdf, "Poster", data.get("poster", []), kol_b)
    else:
        # Flisarbeider
        _pdf_seksjon(pdf, "Flisarbeider", data.get("flis_poster", []), kol_b)

        # Tømrerarbeider
        _pdf_seksjon(pdf, "Tømrerarbeider", data.get("tomrer_poster", []), kol_b)

    # Totaler
    pdf.ln(2)
    tillegg_pst = data.get("tillegg_pst", 0)
    etasje_tillegg = data.get("etasje_tillegg", 0)
    subtotal = data.get("subtotal", 0)
    mva = data.get("mva", 0)
    total_inkl = data.get("total_inkl", 0)

    tom_b = sum(kol_b[:3])  # tomrom foran totaler
    label_w = kol_b[3]
    val_w = kol_b[4]

    pdf.set_font("Helvetica", "", 10)

    if tillegg_pst > 0:
        arbeid = subtotal - etasje_tillegg
        pdf.cell(tom_b, 6, "")
        pdf.cell(label_w, 6, "Sum arbeid:", align="R")
        pdf.cell(val_w, 6, f"{fmt(arbeid)} kr", align="R")
        pdf.ln()
        pdf.cell(tom_b, 6, "")
        etasje = data.get("etasje", 1)
        pdf.cell(label_w, 6, f"Tillegg {tillegg_pst}%:", align="R")
        pdf.cell(val_w, 6, f"{fmt(etasje_tillegg)} kr", align="R")
        pdf.ln()

    pdf.cell(tom_b, 6, "")
    pdf.cell(label_w, 6, "Sum eks. mva:", align="R")
    pdf.cell(val_w, 6, f"{fmt(subtotal)} kr", align="R")
    pdf.ln()

    pdf.cell(tom_b, 6, "")
    pdf.cell(label_w, 6, "MVA 25%:", align="R")
    pdf.cell(val_w, 6, f"{fmt(mva)} kr", align="R")
    pdf.ln()

    pdf.set_draw_color(0, 0, 0)
    pdf.line(10 + tom_b, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(1)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(tom_b, 8, "")
    pdf.cell(label_w, 8, "Total inkl. mva:", align="R")
    pdf.cell(val_w, 8, f"{fmt(total_inkl)} kr", align="R")

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _excel_seksjon(ws, rad, tittel, poster, stiler):
    """poster: liste med (navn, mengde, enhet, enhetspris, total) tupler."""
    if not poster:
        return rad
    ws.cell(row=rad, column=1, value=tittel).font = stiler["seksjon"]
    rad += 1

    for kol, h in enumerate(["Post", "Mengde", "Enhet", "Enh.pris", "Sum kr"], 1):
        c = ws.cell(row=rad, column=kol, value=h)
        c.font = stiler["th_font"]
        c.fill = stiler["th_fill"]
        c.border = stiler["tynn"]
        c.alignment = Alignment(horizontal="center")
    rad += 1

    for navn, mengde, enhet, enhetspris, total in poster:
        ws.cell(row=rad, column=1, value=navn).border = stiler["tynn"]
        c = ws.cell(row=rad, column=2, value=mengde)
        c.border = stiler["tynn"]
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="right")
        ws.cell(row=rad, column=3, value=enhet).border = stiler["tynn"]
        c = ws.cell(row=rad, column=4, value=enhetspris)
        c.border = stiler["tynn"]
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
        c = ws.cell(row=rad, column=5, value=total)
        c.border = stiler["tynn"]
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
        rad += 1

    seksjon_sum = sum(t for _, _, _, _, t in poster)
    c1 = ws.cell(row=rad, column=1, value=f"Sum {tittel.lower()}")
    c1.font = stiler["bold"]
    c1.border = stiler["tynn"]
    c2 = ws.cell(row=rad, column=5, value=seksjon_sum)
    c2.font = stiler["bold"]
    c2.border = stiler["tynn"]
    c2.number_format = "#,##0"
    c2.alignment = Alignment(horizontal="right")
    rad += 2
    return rad


def generer_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kalkyle"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18

    stiler = {
        "header": Font(name="Calibri", bold=True, size=16),
        "tittel": Font(name="Calibri", bold=True, size=18),
        "normal": Font(name="Calibri", size=11),
        "bold": Font(name="Calibri", bold=True, size=11),
        "seksjon": Font(name="Calibri", bold=True, size=13),
        "total": Font(name="Calibri", bold=True, size=14),
        "th_font": Font(name="Calibri", bold=True, color="FFFFFF", size=10),
        "th_fill": PatternFill(start_color="333333", end_color="333333", fill_type="solid"),
        "tynn": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }

    rad = 1
    ws.cell(row=rad, column=1, value=FIRMA["navn"]).font = stiler["header"]
    rad += 1
    ws.cell(row=rad, column=1, value=f"Org.nr: {FIRMA['orgnr']}  |  Tlf: {FIRMA['telefon']}").font = stiler["normal"]
    rad += 1
    ws.cell(row=rad, column=1, value=f"{FIRMA['adresse']}  |  {FIRMA['epost']}").font = stiler["normal"]
    rad += 2

    er_manuell = data.get("kalkyle_type") == "manuell"
    ws.cell(row=rad, column=1, value="KALKYLE" if er_manuell else "BADEROMS KALKYLE").font = stiler["tittel"]
    rad += 2

    if er_manuell:
        info_felter = [
            ("Prosjekt:", data.get("adresse", "")),
            ("Dato:", data.get("dato", "")),
        ]
    else:
        info_felter = [
            ("Prosjekt:", data.get("adresse", "")),
            ("Dato:", data.get("dato", "")),
            ("Gulvareal:", f"{data.get('gulvareal', 0):.2f} m\u00b2"),
            ("Veggareal:", f"{data.get('veggareal', 0):.2f} m\u00b2"),
        ]
    for etikett, verdi in info_felter:
        ws.cell(row=rad, column=1, value=etikett).font = stiler["bold"]
        ws.cell(row=rad, column=2, value=verdi).font = stiler["normal"]
        rad += 1
    rad += 1

    if er_manuell:
        rad = _excel_seksjon(ws, rad, "Poster", data.get("poster", []), stiler)
    else:
        rad = _excel_seksjon(ws, rad, "Flisarbeider", data.get("flis_poster", []), stiler)
        rad = _excel_seksjon(ws, rad, "Tømrerarbeider", data.get("tomrer_poster", []), stiler)

    # Totaler
    tillegg_pst = data.get("tillegg_pst", 0)
    etasje_tillegg = data.get("etasje_tillegg", 0)
    subtotal = data.get("subtotal", 0)
    mva = data.get("mva", 0)
    total_inkl = data.get("total_inkl", 0)

    if tillegg_pst > 0:
        arbeid = subtotal - etasje_tillegg
        ws.cell(row=rad, column=1, value="Sum arbeid:").font = stiler["bold"]
        c = ws.cell(row=rad, column=5, value=arbeid)
        c.font = stiler["bold"]
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
        rad += 1
        etasje = data.get("etasje", 1)
        ws.cell(row=rad, column=4, value=f"Tillegg {tillegg_pst}%:").font = stiler["normal"]
        c = ws.cell(row=rad, column=5, value=etasje_tillegg)
        c.font = stiler["normal"]
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
        rad += 1

    ws.cell(row=rad, column=4, value="Sum eks. mva:").font = stiler["bold"]
    c = ws.cell(row=rad, column=5, value=subtotal)
    c.font = stiler["bold"]
    c.number_format = "#,##0"
    c.alignment = Alignment(horizontal="right")
    rad += 1

    ws.cell(row=rad, column=4, value="MVA 25%:").font = stiler["normal"]
    c = ws.cell(row=rad, column=5, value=mva)
    c.font = stiler["normal"]
    c.number_format = "#,##0"
    c.alignment = Alignment(horizontal="right")
    rad += 1

    ws.cell(row=rad, column=4, value="Total inkl. mva:").font = stiler["total"]
    c = ws.cell(row=rad, column=5, value=total_inkl)
    c.font = stiler["total"]
    c.number_format = "#,##0"
    c.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# E-post
# ---------------------------------------------------------------------------

def send_epost(mottaker, emne, brødtekst, vedlegg, smtp_config):
    """
    Send e-post med vedlegg.

    vedlegg: liste med (filnavn, bytes_data, mime_type) tupler
    smtp_config: dict med host, port, bruker, passord
    """
    msg = MIMEMultipart()
    msg["From"] = smtp_config["bruker"]
    msg["To"] = mottaker
    msg["Subject"] = emne
    msg.attach(MIMEText(brødtekst, "plain", "utf-8"))

    for filnavn, data, mime in vedlegg:
        maintype, subtype = mime.split("/", 1)
        del_obj = MIMEBase(maintype, subtype)
        del_obj.set_payload(data)
        encoders.encode_base64(del_obj)
        del_obj.add_header("Content-Disposition", "attachment", filename=filnavn)
        msg.attach(del_obj)

    with smtplib.SMTP(smtp_config["host"], smtp_config["port"]) as server:
        server.starttls()
        server.login(smtp_config["bruker"], smtp_config["passord"])
        server.send_message(msg)


# ---------------------------------------------------------------------------
# Dokument-PDFer (tekst og bilder)
# ---------------------------------------------------------------------------

def generer_tekst_dokument_pdf(tittel, tekst):
    """Generer en enkel PDF med tittel og brødtekst."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Firmaheader
    pdf.set_font("Helvetica", "B", 16)
    if _LOGO.exists():
        pdf.image(str(_LOGO), x=10, y=10, w=40)
        pdf.set_y(10)
        pdf.cell(45)
    pdf.cell(0, 8, FIRMA["navn"], new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_x(55)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, f"Org.nr: {FIRMA['orgnr']}  |  Tlf: {FIRMA['telefon']}", new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_y(max(pdf.get_y(), 35))
    pdf.ln(3)
    pdf.set_draw_color(180, 180, 180)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(8)

    # Tittel
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, tittel, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Dato
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, datetime.date.today().strftime("%d.%m.%Y"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # Brødtekst
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(0, 6, tekst)

    return bytes(pdf.output())


def generer_kontaktliste_pdf(adresse, kontakter):
    """Generer en PDF med kontaktliste for et prosjekt."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Firmaheader
    pdf.set_font("Helvetica", "B", 16)
    if _LOGO.exists():
        pdf.image(str(_LOGO), x=10, y=10, w=40)
        pdf.set_y(10)
        pdf.cell(45)
    pdf.cell(0, 8, FIRMA["navn"], new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_x(55)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 4, f"Org.nr: {FIRMA['orgnr']}  |  Tlf: {FIRMA['telefon']}", new_x="LMARGIN", new_y="NEXT")
    if _LOGO.exists():
        pdf.set_y(max(pdf.get_y(), 35))
    pdf.ln(3)
    pdf.set_draw_color(180, 180, 180)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(8)

    # Tittel
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Kontaktliste - {adresse}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, datetime.date.today().strftime("%d.%m.%Y"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(6)

    # Kontakter
    for kont in kontakter:
        rolle = kont.get("rolle", "")
        navn = kont.get("navn", "")
        tlf = kont.get("tlf", "")
        epost = kont.get("epost", "")

        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 7, rolle, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 6, f"Navn: {navn}", new_x="LMARGIN", new_y="NEXT")
        if tlf:
            pdf.cell(0, 6, f"Tlf: {tlf}", new_x="LMARGIN", new_y="NEXT")
        if epost:
            pdf.cell(0, 6, f"E-post: {epost}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)
        pdf.set_draw_color(220, 220, 220)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(4)

    if not kontakter:
        pdf.set_font("Helvetica", "I", 11)
        pdf.cell(0, 7, "Ingen kontakter registrert.", new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


def generer_bilde_dokument_pdf(tittel, bilder):
    """Generer en PDF med bilder. bilder: liste med (filnavn, bytes) tupler."""
    from PIL import Image

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for idx, (bilde_navn, bilde_bytes) in enumerate(bilder):
        pdf.add_page()

        if idx == 0:
            # Tittel på første side
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, tittel, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(120, 120, 120)
            pdf.cell(0, 5, datetime.date.today().strftime("%d.%m.%Y"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(4)

        # Last bilde med Pillow for å få dimensjoner
        img = Image.open(io.BytesIO(bilde_bytes))
        img_w, img_h = img.size

        # Tilgjengelig bredde/høyde på siden
        maks_b = 190  # mm (A4 bredde minus marginer)
        maks_h = 260 if idx == 0 else 277  # reduser for tittel på første side

        # Skaler til å passe på siden
        ratio = min(maks_b / (img_w * 0.264583), maks_h / (img_h * 0.264583))
        vis_b = img_w * 0.264583 * ratio  # px til mm
        vis_h = img_h * 0.264583 * ratio

        # Sentrér horisontalt
        x = 10 + (maks_b - vis_b) / 2

        # Skriv bildet til en temp-buffer i et format fpdf2 kan lese
        img_buf = io.BytesIO()
        fmt = "JPEG"
        if img.mode == "RGBA":
            fmt = "PNG"
        img.save(img_buf, format=fmt)
        img_buf.seek(0)

        pdf.image(img_buf, x=x, y=pdf.get_y(), w=vis_b)

        # Bildetekst under bildet
        pdf.set_y(pdf.get_y() + vis_h + 2)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 4, bilde_navn, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)

    return bytes(pdf.output())
